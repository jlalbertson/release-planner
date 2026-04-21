"""Excel (.xlsx) writer: writes candidate and Big Rock data via openpyxl."""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from release_planner.constants import (
    BIG_ROCK_COLUMN_WIDTHS_CHARS,
    BIG_ROCK_COLUMNS,
    FEATURE_COLUMN_WIDTHS_CHARS,
    FEATURE_COLUMNS,
    JIRA_BROWSE_URL,
    RFE_COLUMN_WIDTHS_CHARS,
    RFE_COLUMNS,
    VALIDATION_ISSUE_STATUS,
    VALIDATION_PRIORITY,
    VALIDATION_RFE_STATUS,
)
from release_planner.models import BigRock, Candidate

logger = logging.getLogger(__name__)

# Header styling
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BAND_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_STATUS_DONE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_STATUS_IN_PROGRESS_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_PRIORITY_CRITICAL_FONT = Font(color="FF0000", bold=True)

# Valid Jira issue key pattern
_ISSUE_KEY_RE = re.compile(r"^[A-Z][A-Z0-9]+-\d+$")


class ExcelWriter:
    """Write release planning data to an Excel .xlsx file via openpyxl."""

    def __init__(
        self,
        big_rocks: list[BigRock],
        candidates: dict[str, list[Candidate]],
        release: str,
        fix_versions: list[str] | None = None,
        per_rock_stats: dict[str, dict[str, int]] | None = None,
        outcome_summaries: dict[str, str] | None = None,
    ):
        """Initialize the Excel writer.

        Args:
            big_rocks: Ordered list of BigRock definitions.
            candidates: Map of big_rock_name -> list of Candidate, in display order.
            release: Release version string (e.g. "3.5") for worksheet naming.
            fix_versions: Optional list of target release values for data validation.
            per_rock_stats: Optional dict of rock_name -> {"features": N, "rfes": N}.
            outcome_summaries: Optional dict of outcome_key -> summary (title).
        """
        self._big_rocks = big_rocks
        self._candidates = candidates
        self._release = release
        self._fix_versions = fix_versions or []
        self._per_rock_stats = per_rock_stats or {}
        self._outcome_summaries = outcome_summaries or {}

    def write(self, output_path: str | Path) -> str:
        """Write data to an Excel .xlsx file.

        Creates three worksheets:
        - Proposed Features: RHAISTRAT features only
        - Proposed RFEs: RHAIRFE issues only
        - Big Rocks: rock summary

        Args:
            output_path: Path for the output .xlsx file.

        Returns:
            The absolute path to the written file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Split candidates into features (RHAISTRAT) and RFEs (RHAIRFE)
        features: dict[str, list[Candidate]] = {}
        rfes: dict[str, list[Candidate]] = {}
        for rock_name, candidates in self._candidates.items():
            features[rock_name] = [c for c in candidates if not c.issue_key.startswith("RHAIRFE-")]
            rfes[rock_name] = [c for c in candidates if c.issue_key.startswith("RHAIRFE-")]

        wb = Workbook()

        # Feature worksheet (rename the default sheet)
        ws_features = wb.active
        ws_features.title = "Proposed Features"
        self._write_feature_worksheet(ws_features, features)

        # RFE worksheet
        ws_rfes = wb.create_sheet(title="Proposed RFEs")
        self._write_rfe_worksheet(ws_rfes, rfes)

        # Big Rocks worksheet
        ws_big_rocks = wb.create_sheet(title="Big Rocks")
        self._write_big_rocks_worksheet(ws_big_rocks)

        wb.save(str(output_path))
        abs_path = str(output_path.resolve())
        logger.info("Wrote Excel file: %s", abs_path)
        return abs_path

    def _write_feature_worksheet(self, ws, features: dict[str, list[Candidate]]) -> None:
        """Write the Proposed Features worksheet (RHAISTRAT features only)."""
        columns = FEATURE_COLUMNS
        ws.append(list(columns))

        row_count = 0
        for rock in self._big_rocks:
            for candidate in features.get(rock.name, []):
                ws.append(self._feature_row(candidate))
                row_count += 1
        # Tier 2 features (no Big Rock association)
        for candidate in features.get("", []):
            ws.append(self._feature_row(candidate))
            row_count += 1

        self._format_header(ws, len(columns))
        self._set_column_widths(ws, columns, FEATURE_COLUMN_WIDTHS_CHARS)
        self._apply_conditional_formatting(ws, row_count, columns, "Issue status", "Priority")
        self._apply_row_banding(ws, row_count, len(columns))
        self._apply_hyperlinks_to_column(ws, row_count, columns, "Feature")
        self._apply_hyperlinks_to_column(ws, row_count, columns, "RFE")
        self._apply_data_validations_for(ws, row_count, columns, is_rfe=False)
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{row_count + 1}"

        logger.info("Wrote %d rows to '%s' worksheet", row_count, ws.title)

    def _write_rfe_worksheet(self, ws, rfes: dict[str, list[Candidate]]) -> None:
        """Write the RFE worksheet (RHAIRFE issues only)."""
        columns = RFE_COLUMNS
        ws.append(list(columns))

        row_count = 0
        for rock in self._big_rocks:
            for candidate in rfes.get(rock.name, []):
                ws.append(self._rfe_row(candidate))
                row_count += 1
        # Tier 2 RFEs (no Big Rock association)
        for candidate in rfes.get("", []):
            ws.append(self._rfe_row(candidate))
            row_count += 1

        self._format_header(ws, len(columns))
        self._set_column_widths(ws, columns, RFE_COLUMN_WIDTHS_CHARS)
        self._apply_conditional_formatting(ws, row_count, columns, "RFE Status", "Priority")
        self._apply_row_banding(ws, row_count, len(columns))
        self._apply_hyperlinks_to_column(ws, row_count, columns, "RFE")
        self._apply_data_validations_for(ws, row_count, columns, is_rfe=True)
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{row_count + 1}"

        logger.info("Wrote %d rows to '%s' worksheet", row_count, ws.title)

    def _write_big_rocks_worksheet(self, ws) -> None:
        """Write the 'Big Rocks' worksheet with formatting."""
        # Header row
        ws.append(list(BIG_ROCK_COLUMNS))

        # Data rows
        for rock in self._big_rocks:
            stats = self._per_rock_stats.get(rock.name, {})
            outcome_keys_str = ", ".join(rock.outcome_keys) if rock.outcome_keys else ""
            outcome_desc_parts = [
                self._outcome_summaries.get(k, "") for k in rock.outcome_keys
            ]
            outcome_desc = "; ".join(d for d in outcome_desc_parts if d)
            ws.append(
                [
                    rock.pillar,
                    rock.priority,
                    rock.name,
                    outcome_keys_str,
                    outcome_desc,
                    rock.state,
                    rock.owner,
                    stats.get("features", 0),
                    stats.get("rfes", 0),
                    "",  # Notes left blank for now
                ]
            )

        # Apply formatting
        self._format_header(ws, len(BIG_ROCK_COLUMNS))
        self._set_column_widths(ws, BIG_ROCK_COLUMNS, BIG_ROCK_COLUMN_WIDTHS_CHARS)
        self._merge_pillar_cells(ws, len(self._big_rocks))
        self._apply_outcome_hyperlinks(ws, len(self._big_rocks))
        ws.freeze_panes = "A2"

        logger.info("Wrote %d rows to '%s' worksheet", len(self._big_rocks), ws.title)

    def _feature_row(self, candidate: Candidate) -> list:
        """Convert a Candidate to a row matching FEATURE_COLUMNS order."""
        return [
            candidate.big_rock,
            candidate.issue_key,
            candidate.status,
            candidate.priority,
            candidate.phase,
            candidate.summary,
            candidate.components,
            candidate.target_release,
            candidate.fix_version,
            candidate.pm,
            candidate.delivery_owner,
            candidate.rfe,
            candidate.labels,
        ]

    def _rfe_row(self, candidate: Candidate) -> list:
        """Convert a Candidate to a row matching RFE_COLUMNS order."""
        return [
            candidate.big_rock,
            candidate.issue_key,
            candidate.status,
            candidate.priority,
            candidate.summary,
            candidate.components,
            candidate.pm,
            candidate.labels,
        ]

    @staticmethod
    def _apply_hyperlinks_to_column(
        ws, row_count: int, columns: list[str], col_name: str
    ) -> None:
        """Add hyperlinks to a named column containing Jira issue keys."""
        if col_name not in columns:
            return
        col_idx = columns.index(col_name) + 1
        for row_idx in range(2, row_count + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            key = cell.value
            if key and isinstance(key, str) and _ISSUE_KEY_RE.match(key):
                cell.hyperlink = f"{JIRA_BROWSE_URL}/{key}"
                cell.font = Font(color="0563C1", underline="single")

    @staticmethod
    def _apply_outcome_hyperlinks(ws, row_count: int) -> None:
        """Add hyperlinks to Outcome keys in the Big Rocks worksheet.

        Single keys get a direct hyperlink. Multiple comma-separated keys
        link to a JQL search showing all of them.
        """
        outcome_col = BIG_ROCK_COLUMNS.index("Outcome") + 1
        link_font = Font(color="0563C1", underline="single")
        for row_idx in range(2, row_count + 2):
            cell = ws.cell(row=row_idx, column=outcome_col)
            value = cell.value
            if not value or not isinstance(value, str):
                continue
            keys = [k.strip() for k in value.split(",") if k.strip()]
            if not keys:
                continue
            if all(_ISSUE_KEY_RE.match(k) for k in keys):
                if len(keys) == 1:
                    cell.hyperlink = f"{JIRA_BROWSE_URL}/{keys[0]}"
                else:
                    jql_keys = "%2C+".join(keys)
                    cell.hyperlink = (
                        f"https://redhat.atlassian.net/issues/"
                        f"?jql=key+in+({jql_keys})"
                    )
                cell.font = link_font

    @staticmethod
    def _format_header(ws, num_cols: int) -> None:
        """Apply header row formatting: dark background, white bold text."""
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    @staticmethod
    def _set_column_widths(ws, columns: list[str], widths: dict[str, float]) -> None:
        """Set column widths from character-unit width dict."""
        for i, col_name in enumerate(columns):
            char_width = widths.get(col_name, 13.0)
            col_letter = get_column_letter(i + 1)
            ws.column_dimensions[col_letter].width = char_width

    @staticmethod
    def _apply_conditional_formatting(
        ws, row_count: int, columns: list[str], status_col_name: str, priority_col_name: str
    ) -> None:
        """Apply status and priority formatting to data rows."""
        status_col = columns.index(status_col_name) + 1
        priority_col = columns.index(priority_col_name) + 1

        for row_idx in range(2, row_count + 2):
            # Status formatting
            status_cell = ws.cell(row=row_idx, column=status_col)
            status_val = status_cell.value
            if status_val in ("Done", "Closed"):
                status_cell.fill = _STATUS_DONE_FILL
            elif status_val == "In Progress":
                status_cell.fill = _STATUS_IN_PROGRESS_FILL

            # Priority formatting
            priority_cell = ws.cell(row=row_idx, column=priority_col)
            priority_val = priority_cell.value
            if priority_val in ("Blocker", "Critical"):
                priority_cell.font = _PRIORITY_CRITICAL_FONT

    @staticmethod
    def _apply_row_banding(ws, row_count: int, num_cols: int) -> None:
        """Apply alternating row banding (light gray on even data rows)."""
        for row_idx in range(2, row_count + 2):
            if row_idx % 2 == 1:  # odd data rows (0-indexed: even rows in sheet)
                for col_idx in range(1, num_cols + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if not cell.fill or cell.fill == PatternFill():
                        cell.fill = _BAND_FILL

    def _apply_data_validations_for(
        self, ws, row_count: int, columns: list[str], is_rfe: bool = False
    ) -> None:
        """Add dropdown data validations to columns present in the given column list."""
        if row_count < 1:
            return

        last_row = row_count + 1  # +1 for header offset

        big_rock_names = [rock.name for rock in self._big_rocks]
        validations: list[tuple[str, list[str]]] = [
            ("Big Rock", big_rock_names),
            ("Issue status", VALIDATION_ISSUE_STATUS),
            ("RFE Status", VALIDATION_RFE_STATUS),
            ("Priority", VALIDATION_PRIORITY),
        ]

        if self._fix_versions:
            validations.append(("Target Release", self._fix_versions))

        for col_header, values in validations:
            if col_header not in columns:
                continue
            col_idx = columns.index(col_header) + 1
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{last_row}"

            formula = '"' + ",".join(values) + '"'
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showDropDown=False,
            )
            dv.sqref = cell_range
            ws.add_data_validation(dv)

    @staticmethod
    def _merge_pillar_cells(ws, row_count: int) -> None:
        """Merge contiguous Pillar cells vertically in the Big Rocks worksheet.

        Identifies runs of the same Pillar value in column A (starting at row 2)
        and merges them, applying vertical alignment.
        """
        if row_count < 2:
            return

        pillar_col = 1  # Column A
        start_row = 2  # First data row (after header)

        current_pillar = ws.cell(row=start_row, column=pillar_col).value
        run_start = start_row

        for row_idx in range(start_row + 1, start_row + row_count + 1):
            # Use None for rows past the data to flush the last run
            if row_idx <= start_row + row_count - 1:
                cell_value = ws.cell(row=row_idx, column=pillar_col).value
            else:
                cell_value = None

            if cell_value == current_pillar and row_idx <= start_row + row_count - 1:
                continue

            # End of a run -- merge if more than 1 row
            run_end = row_idx - 1
            if run_end > run_start:
                ws.merge_cells(
                    start_row=run_start,
                    start_column=pillar_col,
                    end_row=run_end,
                    end_column=pillar_col,
                )
                ws.cell(row=run_start, column=pillar_col).alignment = Alignment(
                    vertical="center",
                )

            # Start new run
            if row_idx <= start_row + row_count - 1:
                current_pillar = cell_value
                run_start = row_idx
