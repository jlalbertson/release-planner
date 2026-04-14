"""Spreadsheet import: reads .xlsx files and generates overrides YAML.

Uses openpyxl, which is an optional dependency (pip install release-planner[import]).
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import yaml

logger = logging.getLogger(__name__)

# Known column name variations (lowercased) -> canonical field name
COLUMN_ALIASES: dict[str, str] = {
    "big rock": "big_rock",
    "big rocks": "big_rock",
    "1-n ranking": "ranking",
    "ranking": "ranking",
    "feature": "issue_key",
    "issue key": "issue_key",
    "issue_key": "issue_key",
    "key": "issue_key",
    "issue status": "status",
    "status": "status",
    "priority": "priority",
    "dp/tp/ga": "phase",
    "phase": "phase",
    "title": "summary",
    "summary": "summary",
    "team": "team",
    "component[s]": "components",
    "components": "components",
    "component": "components",
    "target release": "target_release",
    "target_release": "target_release",
    "rfe": "rfe",
    "rfe status": "rfe_status",
    "rfe_status": "rfe_status",
    "pm": "pm",
    "architect": "architect",
    "delivery owner": "delivery_owner",
    "delivery_owner": "delivery_owner",
    "risk flag": "risk_flag",
    "risk_flag": "risk_flag",
    "change log": "change_log",
    "change_log": "change_log",
    "refinement complete": "refinement_complete",
    "refinement_complete": "refinement_complete",
    "refinement notes": "refinement_notes",
    "refinement_notes": "refinement_notes",
    "comments": "comments",
    "rice score": "rice_score",
    "rice_score": "rice_score",
}


class SpreadsheetImporter:
    """Import an existing .xlsx spreadsheet to bootstrap overrides.yaml."""

    def __init__(self, xlsx_path: str, sheet_name: str | None = None):
        """Load the workbook via openpyxl.

        Auto-detects sheet name if not specified (looks for sheets matching
        '*Candidates*' or '*Commitments*').

        Args:
            xlsx_path: Path to the .xlsx file.
            sheet_name: Specific sheet name to import. Auto-detects if None.

        Raises:
            ImportError: If openpyxl is not installed.
            FileNotFoundError: If the .xlsx file does not exist.
        """
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise ImportError(
                "openpyxl is required for the import-xlsx command. "
                "Install it with: pip install release-planner[import]"
            ) from None

        path = Path(xlsx_path)
        if not path.exists():
            raise FileNotFoundError(f"Spreadsheet not found: {xlsx_path}")

        import openpyxl

        self._wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        self._sheet_name = sheet_name
        self._ws = self._detect_worksheet()

    def _detect_worksheet(self):
        """Find the worksheet to import."""
        if self._sheet_name:
            if self._sheet_name in self._wb.sheetnames:
                return self._wb[self._sheet_name]
            raise ValueError(
                f"Sheet '{self._sheet_name}' not found. "
                f"Available sheets: {', '.join(self._wb.sheetnames)}"
            )

        # Auto-detect: look for sheets matching known patterns
        for name in self._wb.sheetnames:
            if "candidates" in name.lower() or "commitments" in name.lower():
                logger.info("Auto-detected sheet: %s", name)
                return self._wb[name]

        # Fall back to first sheet
        first = self._wb.sheetnames[0]
        logger.info("Using first sheet: %s", first)
        return self._wb[first]

    def import_to_overrides(self, output_path: str) -> int:
        """Read all rows, extract issue keys and field values, write as overrides YAML.

        For rows with an issue key (RHAISTRAT-xxx, RHOAIENG-xxx, RHAIRFE-xxx):
            Uses the issue key as the override key.
        For rows with no issue key but an RFE key:
            Uses the RFE key as the override key.
        For rows with neither:
            Generates a MANUAL-NNN key.

        Args:
            output_path: Path to write the overrides YAML file.

        Returns:
            Count of entries imported.
        """
        rows = list(self._ws.iter_rows(values_only=True))
        if not rows:
            logger.warning("Spreadsheet has no data")
            return 0

        # First row is header
        header = [str(cell).strip() if cell else "" for cell in rows[0]]
        col_map = self._detect_columns(header)

        if not col_map:
            logger.error("Could not detect any known columns in header: %s", header)
            return 0

        logger.info("Detected columns: %s", col_map)

        overrides: dict[str, dict[str, str | int | float | None]] = {}
        manual_counter = 1

        for row_idx, row in enumerate(rows[1:], start=2):
            # Build field dict for this row
            fields: dict[str, str | int | float | None] = {}
            for canonical_name, col_idx in col_map.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if value is not None:
                        if isinstance(value, (int, float)):
                            fields[canonical_name] = value
                        else:
                            fields[canonical_name] = str(value).strip()

            # Skip empty rows
            if not fields or all(v == "" for v in fields.values()):
                continue

            # Determine the key for this override
            issue_key = self._extract_key(fields.get("issue_key", ""))
            rfe_key = self._extract_key(fields.get("rfe", ""))

            if issue_key:
                key = issue_key
                # Remove issue_key from override fields (it IS the key)
                fields.pop("issue_key", None)
            elif rfe_key:
                key = rfe_key
                fields.pop("rfe", None)
            else:
                key = f"MANUAL-{manual_counter:03d}"
                manual_counter += 1
                # Mark as manual entry
                fields["source"] = "manual"

            # Remove empty string values (no point in overriding with empty)
            fields = {k: v for k, v in fields.items() if v != "" and v is not None}

            if fields:
                overrides[key] = fields

        # Write YAML
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        with open(output, "w") as f:
            f.write("# Auto-generated overrides from spreadsheet import\n")
            f.write(f"# Source: {self._ws.title}\n")
            f.write(f"# Entries: {len(overrides)}\n\n")
            yaml.dump(overrides, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

        logger.info("Wrote %d overrides to %s", len(overrides), output_path)
        return len(overrides)

    def _detect_columns(self, header_row: list[str]) -> dict[str, int]:
        """Map known column names to column indices. Handles minor variations.

        Args:
            header_row: List of header cell values.

        Returns:
            Dict mapping canonical field name to column index.
        """
        col_map: dict[str, int] = {}

        for idx, header in enumerate(header_row):
            if not header:
                continue
            normalized = header.lower().strip()
            canonical = COLUMN_ALIASES.get(normalized)
            if canonical and canonical not in col_map:
                col_map[canonical] = idx

        return col_map

    @staticmethod
    def _extract_key(value: str | int | float | None) -> str:
        """Extract a Jira issue key from a cell value.

        Handles:
        - Direct keys: RHOAIENG-12345
        - URLs: https://issues.redhat.com/browse/RHOAIENG-12345
        - URLs: https://redhat.atlassian.net/browse/RHAISTRAT-9066

        Args:
            value: Cell value that may contain an issue key.

        Returns:
            Extracted issue key, or empty string.
        """
        if not value or value == "":
            return ""
        text = str(value).strip()
        if not text:
            return ""

        # Check for URL-format keys
        url_match = re.match(r"https?://.+/browse/([A-Z]+-\d+)", text)
        if url_match:
            return url_match.group(1)

        # Check for direct issue keys
        key_match = re.match(r"^([A-Z]+-\d+)$", text)
        if key_match:
            return key_match.group(1)

        return ""
