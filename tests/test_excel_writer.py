"""Tests for the Excel (.xlsx) writer."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from release_planner.constants import BIG_ROCK_COLUMNS, FEATURE_COLUMNS, RFE_COLUMNS
from release_planner.excel_writer import ExcelWriter
from release_planner.models import BigRock, Candidate


@pytest.fixture
def simple_big_rocks() -> list[BigRock]:
    """Two BigRocks for testing."""
    return [
        BigRock(
            priority=1,
            name="MaaS",
            full_name="MaaS (continue from 3.4)",
            pillar="Inference",
            state="continue from 3.4",
            outcome_keys=["RHAISTRAT-9001"],
        ),
        BigRock(
            priority=2,
            name="Gen AI Studio",
            full_name="Gen AI Studio",
            pillar="Agents",
            state="new for 3.5",
            outcome_keys=["RHAISTRAT-9002"],
        ),
    ]


@pytest.fixture
def simple_candidates() -> dict[str, list[Candidate]]:
    """Candidates mapped by rock name -- mix of features and RFEs."""
    return {
        "MaaS": [
            Candidate(
                big_rock="MaaS",
                issue_key="RHAISTRAT-100",
                status="In Progress",
                priority="Major",
                summary="Feature A",
                labels="rhoai-3.5, 3.5-candidate",
                source_pass="outcome",
            ),
            Candidate(
                big_rock="MaaS",
                issue_key="RHAISTRAT-101",
                status="Closed",
                priority="Critical",
                summary="Feature B",
                source_pass="outcome",
            ),
            Candidate(
                big_rock="MaaS",
                issue_key="RHAIRFE-50",
                status="Approved",
                priority="Major",
                summary="RFE item A",
                source_pass="outcome",
            ),
        ],
        "Gen AI Studio": [
            Candidate(
                big_rock="Gen AI Studio",
                issue_key="RHAISTRAT-200",
                status="New",
                priority="Blocker",
                summary="Feature C",
                source_pass="outcome",
            ),
            Candidate(
                big_rock="Gen AI Studio",
                issue_key="RHAIRFE-60",
                status="New",
                priority="Normal",
                summary="RFE item B",
                source_pass="outcome",
            ),
        ],
    }


class TestExcelWriter:
    """Test ExcelWriter output."""

    def test_write_creates_file(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        out = tmp_path / "test.xlsx"
        result = writer.write(out)
        assert Path(result).exists()

    def test_write_returns_absolute_path(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        out = tmp_path / "test.xlsx"
        result = writer.write(out)
        assert Path(result).is_absolute()

    def test_creates_parent_directories(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        out = tmp_path / "nested" / "dir" / "test.xlsx"
        result = writer.write(out)
        assert Path(result).exists()

    def test_three_worksheets_created(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        assert len(wb.sheetnames) == 3
        assert "Proposed Features" in wb.sheetnames
        assert "Proposed RFEs" in wb.sheetnames
        assert "Big Rocks" in wb.sheetnames

    def test_candidates_worksheet_names_are_static(
        self, tmp_path, simple_big_rocks, simple_candidates
    ):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "4.0")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        assert "Proposed Features" in wb.sheetnames
        assert "Proposed RFEs" in wb.sheetnames


class TestFeatureWorksheet:
    """Test the Engineering Commitments (Feature) worksheet."""

    def test_correct_column_count(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        header_row = [cell.value for cell in ws[1] if cell.value is not None]
        assert len(header_row) == len(FEATURE_COLUMNS)

    def test_correct_header_order(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(FEATURE_COLUMNS))]
        assert headers == FEATURE_COLUMNS

    def test_only_features_not_rfes(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        # Should have 3 features (RHAISTRAT-100, 101, 200), no RHAIRFE
        data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
        assert len(data_rows) == 3
        feature_col = FEATURE_COLUMNS.index("Feature")
        keys = [row[feature_col] for row in data_rows]
        assert all(k.startswith("RHAISTRAT-") for k in keys)

    def test_data_values(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        assert ws.cell(row=2, column=1).value == "MaaS"
        assert ws.cell(row=2, column=2).value == "RHAISTRAT-100"

    def test_feature_hyperlinks(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        cell = ws.cell(row=2, column=2)  # Feature column
        assert cell.hyperlink is not None
        assert "RHAISTRAT-100" in cell.hyperlink.target

    def test_frozen_header_row(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        assert ws.freeze_panes == "C2"

    def test_auto_filter(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        assert ws.auto_filter.ref is not None

    def test_header_formatting(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True

    def test_status_done_formatting(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        status_col = FEATURE_COLUMNS.index("Issue status") + 1
        # Row 3 has status "Closed"
        cell = ws.cell(row=3, column=status_col)
        assert cell.fill.start_color.rgb == "00C6EFCE"

    def test_priority_critical_formatting(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        priority_col = FEATURE_COLUMNS.index("Priority") + 1
        # Row 3 has priority "Critical"
        cell = ws.cell(row=3, column=priority_col)
        assert cell.font.bold is True

    def test_labels_in_comments(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        comments_col = FEATURE_COLUMNS.index("Comments") + 1
        cell = ws.cell(row=2, column=comments_col)
        assert "rhoai-3.5" in str(cell.value)
        assert "3.5-candidate" in str(cell.value)

    def test_has_phase_column(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(FEATURE_COLUMNS))]
        assert "DP/TP/GA" in headers

    def test_no_team_column(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(FEATURE_COLUMNS))]
        assert "Team" not in headers

    def test_no_rfe_status_column(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(FEATURE_COLUMNS))]
        assert "RFE" in headers
        assert "RFE Status" not in headers


class TestRFEWorksheet:
    """Test the RFE worksheet."""

    def test_correct_column_count(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed RFEs"]
        header_row = [cell.value for cell in ws[1] if cell.value is not None]
        assert len(header_row) == len(RFE_COLUMNS)

    def test_correct_header_order(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed RFEs"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(RFE_COLUMNS))]
        assert headers == RFE_COLUMNS

    def test_only_rfes_not_features(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed RFEs"]
        # Should have 2 RFEs (RHAIRFE-50, RHAIRFE-60)
        data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row[0]]
        assert len(data_rows) == 2
        rfe_col = RFE_COLUMNS.index("RFE")
        keys = [row[rfe_col] for row in data_rows]
        assert all(k.startswith("RHAIRFE-") for k in keys)

    def test_rfe_hyperlinks(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed RFEs"]
        cell = ws.cell(row=2, column=2)  # RFE column
        assert cell.hyperlink is not None
        assert "RHAIRFE-50" in cell.hyperlink.target

    def test_no_team_column(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed RFEs"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(RFE_COLUMNS))]
        assert "Team" not in headers

    def test_no_delivery_owner_column(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed RFEs"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(RFE_COLUMNS))]
        assert "Delivery Owner" not in headers

    def test_no_phase_column(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed RFEs"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(RFE_COLUMNS))]
        assert "DP/TP/GA" not in headers

    def test_rfe_status_is_column_c(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed RFEs"]
        assert ws.cell(row=1, column=3).value == "RFE Status"


class TestBigRocksWorksheet:
    """Test the Big Rocks worksheet content."""

    def test_correct_column_count(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        header_row = [cell.value for cell in ws[1] if cell.value is not None]
        assert len(header_row) == len(BIG_ROCK_COLUMNS)

    def test_correct_header_order(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        headers = [ws.cell(row=1, column=i + 1).value for i in range(len(BIG_ROCK_COLUMNS))]
        assert headers == BIG_ROCK_COLUMNS

    def test_correct_data_row_count(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        data_rows = [row for row in ws.iter_rows(min_row=2) if row[0].value is not None]
        assert len(data_rows) == 2

    def test_notes_column_is_blank(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        notes_col = BIG_ROCK_COLUMNS.index("Notes") + 1
        notes = ws.cell(row=2, column=notes_col).value
        assert notes is None or notes == ""

    def test_frozen_header_row(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        assert ws.freeze_panes == "A2"


class TestDataValidations:
    """Test data validation dropdowns on the Feature worksheet."""

    def test_data_validations_present(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        assert len(ws.data_validations.dataValidation) >= 3

    def test_big_rock_validation_values(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        big_rock_dv = None
        for dv in ws.data_validations.dataValidation:
            if "A2:" in str(dv.sqref):
                big_rock_dv = dv
                break
        assert big_rock_dv is not None
        assert "MaaS" in big_rock_dv.formula1
        assert "Gen AI Studio" in big_rock_dv.formula1

    def test_issue_status_validation_values(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        status_col = FEATURE_COLUMNS.index("Issue status") + 1
        col_letter = chr(ord("A") + status_col - 1)
        status_dv = None
        for dv in ws.data_validations.dataValidation:
            if f"{col_letter}2:" in str(dv.sqref):
                status_dv = dv
                break
        assert status_dv is not None
        assert "In Progress" in status_dv.formula1

    def test_priority_validation_values(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        priority_col = FEATURE_COLUMNS.index("Priority") + 1
        col_letter = chr(ord("A") + priority_col - 1)
        priority_dv = None
        for dv in ws.data_validations.dataValidation:
            if f"{col_letter}2:" in str(dv.sqref):
                priority_dv = dv
                break
        assert priority_dv is not None
        assert "Blocker" in priority_dv.formula1

    def test_no_validations_on_empty_data(self, tmp_path):
        rocks = [
            BigRock(
                priority=1,
                name="EmptyRock",
                full_name="Empty Rock",
                outcome_keys=[],
            ),
        ]
        candidates: dict[str, list[Candidate]] = {"EmptyRock": []}
        writer = ExcelWriter(rocks, candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        assert len(ws.data_validations.dataValidation) == 0


class TestPillarMerging:
    """Test Pillar cell merging in the Summit Big Rocks worksheet."""

    @pytest.fixture
    def multi_pillar_rocks(self) -> list[BigRock]:
        """Multiple rocks with shared pillars for merge testing."""
        return [
            BigRock(priority=1, name="MaaS", full_name="MaaS", pillar="Inference",
                    outcome_keys=["RHAISTRAT-9001"]),
            BigRock(priority=2, name="Tool Calling", full_name="Tool Calling", pillar="Inference",
                    outcome_keys=[]),
            BigRock(priority=3, name="vLLM", full_name="vLLM", pillar="Inference",
                    outcome_keys=[]),
            BigRock(priority=4, name="Gen AI Studio", full_name="Gen AI Studio", pillar="Agents",
                    outcome_keys=["RHAISTRAT-9002"]),
            BigRock(priority=5, name="BYO Agent", full_name="BYO Agent", pillar="Agents",
                    outcome_keys=[]),
            BigRock(priority=6, name="AutoRAG", full_name="AutoRAG", pillar="Data",
                    outcome_keys=["RHAISTRAT-9012"]),
        ]

    def test_pillar_cells_are_merged(self, tmp_path, multi_pillar_rocks):
        candidates: dict[str, list[Candidate]] = {}
        writer = ExcelWriter(multi_pillar_rocks, candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        merged = [str(r) for r in ws.merged_cells.ranges]
        assert "A2:A4" in merged
        assert "A5:A6" in merged

    def test_single_pillar_not_merged(self, tmp_path, multi_pillar_rocks):
        candidates: dict[str, list[Candidate]] = {}
        writer = ExcelWriter(multi_pillar_rocks, candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        merged = [str(r) for r in ws.merged_cells.ranges]
        assert "A7:A7" not in merged

    def test_merged_cell_alignment(self, tmp_path, multi_pillar_rocks):
        candidates: dict[str, list[Candidate]] = {}
        writer = ExcelWriter(multi_pillar_rocks, candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        cell = ws.cell(row=2, column=1)
        assert cell.alignment.vertical == "center"

    def test_no_merge_with_single_rock(self, tmp_path, simple_big_rocks, simple_candidates):
        writer = ExcelWriter(simple_big_rocks, simple_candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Big Rocks"]
        merged = [str(r) for r in ws.merged_cells.ranges]
        assert len(merged) == 0


class TestEmptyRock:
    """Test handling of a rock with no candidates."""

    def test_empty_rock_handled_gracefully(self, tmp_path):
        rocks = [
            BigRock(
                priority=1,
                name="EmptyRock",
                full_name="Empty Rock",
                outcome_keys=[],
            ),
        ]
        candidates: dict[str, list[Candidate]] = {"EmptyRock": []}
        writer = ExcelWriter(rocks, candidates, "3.5")
        writer.write(tmp_path / "test.xlsx")
        wb = load_workbook(tmp_path / "test.xlsx")
        ws = wb["Proposed Features"]
        data_rows = [row for row in ws.iter_rows(min_row=2) if row[0].value is not None]
        assert len(data_rows) == 0
